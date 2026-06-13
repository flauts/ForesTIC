from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
OUTPUT = DATA_DIR / "real_forest.json"


def main() -> None:
    censo_path = next(DATA_DIR.rglob("BD - CENSO FORESTAL.xlsx"))
    muestra_path = next(DATA_DIR.rglob("BD - MUESTRA SUPERVISADA.xlsx"))
    loe_paths = sorted(
        path
        for path in DATA_DIR.rglob("*.xlsx")
        if "talado" in path.name.lower() or "trozado" in path.name.lower()
    )

    censo_rows = read_table(censo_path, "CENSO", header_row=1)
    muestra_rows = read_table(muestra_path, "Hoja1", header_row=1)

    censo = build_censo(censo_rows)
    supervision = build_supervision(muestra_rows)
    loe_payload = build_loe(loe_paths)
    balances = build_balances()
    alertas = build_alertas(balances)

    first_censo = censo_rows[0]
    raw = {
        "metadata": {
            "dataset": "REAL_FILES_CANONICAL_V1",
            "generated_at": datetime.now(UTC).isoformat(timespec="seconds"),
            "source_files": [
                str(path.relative_to(ROOT)).replace("\\", "/")
                for path in [censo_path, muestra_path, *loe_paths, *DATA_DIR.rglob("*.pdf")]
            ],
        },
        "titulos_habilitantes": [
            {
                "titulo_habilitante_id": "TH-BELGICA-REAL",
                "codigo_titulo": clean(first_censo.get("NUM_THABILITANTE")),
                "tipo_titulo": clean(first_censo.get("MODALIDAD")),
                "region": clean(first_censo.get("DEPARTAMENTO")),
                "estado": "vigente",
                "fecha_inicio": "2025-07-22",
                "fecha_fin": "2028-07-22",
                "fuente_origen": "ARCHIVOS_COMPETENCIA",
            }
        ],
        "planes_operativos": [
            {
                "plan_operativo_id": "POA-BELGICA-REAL",
                "titulo_habilitante_id": "TH-BELGICA-REAL",
                "codigo_poa": clean(first_censo.get("NOMBRE_POA")),
                "periodo_inicio": "2025-07-22",
                "periodo_fin": "2028-07-22",
                "estado_aprobacion": "aprobado",
                "volumen_total_autorizado_m3": round(sum(float(row["volumen_estimado_m3"]) for row in censo), 3),
                "fuente_origen": "ARCHIVOS_COMPETENCIA",
            }
        ],
        "parcelas_corta": build_parcelas(censo_rows),
        "censo_forestal": censo,
        "muestra_supervisada": supervision,
        "trozas": loe_payload["trozas"],
        "evento_loe": loe_payload["eventos"],
        "gtf": loe_payload["gtf"],
        "balance_extraccion": balances,
        "alertas": alertas,
    }

    OUTPUT.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"saved {OUTPUT}")
    print(
        json.dumps(
            {
                "arboles": len(raw["censo_forestal"]),
                "muestras": len(raw["muestra_supervisada"]),
                "trozas": len(raw["trozas"]),
                "gtf": len(raw["gtf"]),
                "balances": len(raw["balance_extraccion"]),
                "alertas": len(raw["alertas"]),
            },
            ensure_ascii=False,
        )
    )


def read_table(path: Path, sheet_name: str, header_row: int) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers: list[str] = []
    data: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        if idx == header_row:
            headers = [clean(value) for value in row]
            continue
        if idx > header_row and any(value is not None for value in row):
            data.append({headers[pos]: value for pos, value in enumerate(row) if pos < len(headers) and headers[pos]})
    wb.close()
    return data


def read_loe_table(path: Path, sheet_name: str) -> tuple[dict[str, str], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    metadata: dict[str, str] = {}
    rows = ws.iter_rows(values_only=True)
    headers: list[str] | None = None
    data: list[dict[str, Any]] = []
    for row in rows:
        label = clean(row[0]) if row and row[0] is not None else ""
        if label in {"Titular", "Titulo Habilitante", "N° Resolucion", "Vigencia"}:
            metadata[label] = clean(next((value for value in row[1:] if value is not None), ""))
        second = clean(row[1]).lower() if len(row) > 1 and row[1] is not None else ""
        if second in {"fecha", "fechas"}:
            headers = [clean(value) for value in row]
            break
    if headers:
        for row in rows:
            if row and row[0] is not None:
                data.append({headers[pos]: value for pos, value in enumerate(row) if pos < len(headers) and headers[pos]})
    wb.close()
    return metadata, data


def build_censo(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        pca = canonical_pc(row.get("PCA"))
        codigo = clean(row.get("CODIGO"))
        out.append(
            {
                "cod_arbol": cod_arbol(pca, codigo),
                "codigo_original": codigo,
                "parcela_corta_id": pca,
                "especie_declarada": clean(row.get("ESPECIES")),
                "coordenada_utm_este": to_float(row.get("COORDENADA_ESTE")),
                "coordenada_utm_norte": to_float(row.get("COORDENADA_NORTE")),
                "dap_cm": to_float(row.get("DAP")),
                "altura_m": to_float(row.get("AC")),
                "volumen_estimado_m3": to_float(row.get("VOLUMEN")),
                "condicion_censal": clean(row.get("CONDICION")),
                "estado_censal": clean(row.get("ESTADO")),
                "fuente_origen": "CENSO_FORESTAL_REAL",
            }
        )
    return out


def build_supervision(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for idx, row in enumerate(rows, start=1):
        pca = canonical_pc(row.get("PCA_CAMPO") or row.get("PCA"))
        codigo = clean(row.get("CODIGO_CAMPO") or row.get("CODIGO"))
        este = to_float(row.get("COORDENADA_ESTE"))
        norte = to_float(row.get("COORDENADA_NORTE"))
        este_campo = to_float(row.get("COORDENADA_ESTE_CAMPO"))
        norte_campo = to_float(row.get("COORDENADA_NORTE_CAMPO"))
        out.append(
            {
                "muestra_id": f"MS-REAL-{idx:04d}",
                "cod_arbol": cod_arbol(pca, codigo),
                "supervision_id": clean(row.get("NUM_INFORME")),
                "estado_supervision": normalize_supervision_state(row),
                "estado_campo": clean(row.get("DESC_EESTADO_CAMPO")),
                "condicion_campo": clean(row.get("DESC_ECONDICION_CAMPO")),
                "especie_verificada": clean(row.get("DESC_ESPECIES_CAMPO")),
                "utm_este_verificado": este_campo,
                "utm_norte_verificado": norte_campo,
                "desviacion_metros": distance(este, norte, este_campo, norte_campo),
                "fecha_supervision": iso_date(row.get("FECHA_INFORME")),
                "evidencia_hash": sha256_json(row),
                "archivo_geomatica_ref": "BD - MUESTRA SUPERVISADA.xlsx",
            }
        )
    return out


def build_loe(paths: list[Path]) -> dict[str, list[dict[str, Any]]]:
    trozado_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    eventos: list[dict[str, Any]] = []
    despacho_rows: list[dict[str, Any]] = []

    for path in paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = {name.lower(): name for name in wb.sheetnames}
        wb.close()
        metadata, tala = read_loe_table(path, names["tala"])
        pc = pc_from_metadata(metadata, path)
        for row in tala:
            codigo = clean(row.get("Codigo del Arbol"))
            eventos.append(
                {
                    "evento_loe_id": f"LOE-TALA-{pc}-{codigo}-{len(eventos)+1}",
                    "tipo_evento": "tala",
                    "cod_arbol": cod_arbol(pc, codigo),
                    "troza_id": None,
                    "fecha_evento": iso_date(row.get("Fecha")),
                    "volumen_evento_m3": to_float(row.get("Volumen(m3)")),
                    "usuario_origen_anonimizado": "LOE_REAL",
                    "fuente_origen": path.name,
                    "payload_hash": sha256_json(row),
                }
            )

        _, trozado = read_loe_table(path, names["trozado"])
        for row in trozado:
            codigo_troza = clean(row.get("Codigo de la Troza"))
            if not codigo_troza:
                continue
            parent = codigo_troza.split("/")[0].split("-")[0]
            troza_id = make_troza_id(pc, codigo_troza)
            troza = {
                "troza_id": troza_id,
                "codigo_troza": codigo_troza,
                "cod_arbol": cod_arbol(pc, parent),
                "especie": clean(row.get("Especie")),
                "volumen_m3": to_float(row.get("Volumen(m3)")),
                "longitud_m": to_float(row.get("Longitud aprovechable(m)")),
                "diametro_promedio_cm": round(
                    ((to_float(row.get("Diametro Mayor(m)")) + to_float(row.get("Diametro Menor(m)"))) / 2) * 100,
                    2,
                ),
                "estado": "trozada",
                "fuente_origen": path.name,
            }
            trozado_by_key[(pc, codigo_troza)] = troza
            eventos.append(
                {
                    "evento_loe_id": f"LOE-TROZADO-{troza_id}",
                    "tipo_evento": "trozado",
                    "cod_arbol": troza["cod_arbol"],
                    "troza_id": troza_id,
                    "fecha_evento": iso_date(row.get("Fecha")),
                    "volumen_evento_m3": troza["volumen_m3"],
                    "usuario_origen_anonimizado": "LOE_REAL",
                    "fuente_origen": path.name,
                    "payload_hash": sha256_json(row),
                }
            )

        _, despacho = read_loe_table(path, names["despacho"])
        for row in despacho:
            row["_pc"] = pc
            row["_source_file"] = path.name
            despacho_rows.append(row)

    dispatched_by_gtf: dict[str, list[dict[str, Any]]] = defaultdict(list)
    trozas: dict[str, dict[str, Any]] = {}
    for row in despacho_rows:
        pc = row["_pc"]
        codigo_troza = clean(row.get("Codigo de Troza"))
        numero_gtf = clean(row.get("Numero de GTF"))
        if not codigo_troza or not numero_gtf:
            continue
        troza = trozado_by_key.get((pc, codigo_troza))
        if not troza:
            continue
        troza = {**troza, "estado": "despachada"}
        trozas[troza["troza_id"]] = troza
        dispatched_by_gtf[numero_gtf].append(troza)
        eventos.append(
            {
                "evento_loe_id": f"LOE-DESPACHO-{make_troza_id(pc, codigo_troza)}",
                "tipo_evento": "despacho",
                "cod_arbol": troza["cod_arbol"],
                "troza_id": troza["troza_id"],
                "fecha_evento": iso_date(row.get("Fechas")),
                "volumen_evento_m3": troza["volumen_m3"],
                "usuario_origen_anonimizado": "LOE_REAL",
                "fuente_origen": row["_source_file"],
                "payload_hash": sha256_json(row),
            }
        )

    gtf = []
    for numero_gtf, gtf_trozas in sorted(dispatched_by_gtf.items()):
        species = sorted({troza["especie"] for troza in gtf_trozas if troza["especie"]})
        pcs = sorted({troza["cod_arbol"].split("-", 2)[0] + "-" + troza["cod_arbol"].split("-", 2)[1] for troza in gtf_trozas})
        gtf.append(
            {
                "gtf_id": f"GTF-{numero_gtf}",
                "numero_gtf": numero_gtf,
                "fecha_emision": "2025-07-22",
                "fecha_vencimiento": "2028-07-22",
                "origen": f"Comunidad Nativa Belgica - {', '.join(pcs)}",
                "destino": "Destino registrado en LOE",
                "volumen_total_m3": round(sum(float(troza["volumen_m3"]) for troza in gtf_trozas), 3),
                "estado_gtf": "vigente",
                "lote_id": f"LOTE-{numero_gtf}",
                "especie": species[0] if len(species) == 1 else "Mixto",
                "especies": species,
                "trozas": [troza["troza_id"] for troza in gtf_trozas],
                "payload_hash": sha256_json({"numero_gtf": numero_gtf, "trozas": [t["troza_id"] for t in gtf_trozas]}),
            }
        )
    return {"trozas": list(trozas.values()), "eventos": eventos, "gtf": gtf}


def build_balances() -> list[dict[str, Any]]:
    balances: list[dict[str, Any]] = []
    line_re = re.compile(r"^(MADERA EN ROLLO|RAMAS PARA LEÑA)\s+Metros Cúbicos\s+(.+?)\s+(-?\d+\.\d{3})\s+(-?\d+\.\d{3})\s+(-?\d+\.\d{3})$")
    for path in DATA_DIR.rglob("*.pdf"):
        text = "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)
        pc = canonical_pc(re.search(r"\(PC\s*(\d+)\)", text).group(1) if re.search(r"\(PC\s*(\d+)\)", text) else path.stem)
        fecha_match = re.search(r"Fecha de consulta:\s*([0-9/]+)", text)
        for line in text.splitlines():
            match = line_re.match(" ".join(line.split()))
            if not match:
                continue
            producto, especie, autorizado, extraido, saldo = match.groups()
            balances.append(
                {
                    "balance_id": f"BAL-{pc}-{len(balances)+1:03d}",
                    "plan_operativo_id": "POA-BELGICA-REAL",
                    "parcela_corta_id": pc,
                    "producto": producto,
                    "especie": especie.strip(),
                    "volumen_autorizado_m3": float(autorizado),
                    "volumen_movilizado_m3": float(extraido),
                    "volumen_disponible_m3": float(saldo),
                    "fecha_corte": parse_slash_date(fecha_match.group(1)) if fecha_match else "2026-06-10",
                    "fuente_origen": path.name,
                }
            )
    return balances


def build_alertas(balances: list[dict[str, Any]]) -> list[dict[str, Any]]:
    alertas = []
    for row in balances:
        if row["volumen_disponible_m3"] < 0:
            alertas.append(
                {
                    "alerta_id": f"AL-BAL-{len(alertas)+1:04d}",
                    "ambito": "parcela",
                    "referencia_id": row["parcela_corta_id"],
                    "tipo_alerta": "saldo_negativo",
                    "severidad": "critica",
                    "estado_alerta": "vigente",
                    "fecha_alerta": row["fecha_corte"],
                    "descripcion_normalizada": f"Saldo negativo en {row['parcela_corta_id']} para {row['especie']}.",
                }
            )
    return alertas


def build_parcelas(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_pc: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_pc[canonical_pc(row.get("PCA"))].append(row)
    return [
        {
            "parcela_corta_id": pc,
            "plan_operativo_id": "POA-BELGICA-REAL",
            "codigo_pc": pc,
            "area_ha": {"PC-01": 1400.0, "PC-02": 2699.4, "PC-03": 2691.2}.get(pc, 0.0),
            "volumen_autorizado_m3": round(sum(to_float(row.get("VOLUMEN")) for row in rows_pc), 3),
            "estado": "autorizada",
        }
        for pc, rows_pc in sorted(by_pc.items())
    ]


def normalize_supervision_state(row: dict[str, Any]) -> str:
    estado = clean(row.get("DESC_EESTADO_CAMPO")).lower()
    coincide = clean(row.get("COINCIDE_ESPECIES")).upper()
    if "no ubicado" in estado or "inexist" in estado:
        return "inexistente"
    if coincide == "NO":
        return "observado"
    return "conforme"


def canonical_pc(value: Any) -> str:
    text = clean(value).upper()
    match = re.search(r"PC\s*0?(\d+)", text)
    if match:
        return f"PC-{int(match.group(1)):02d}"
    match = re.search(r"\b0?([123])\b", text)
    return f"PC-{int(match.group(1)):02d}" if match else "PC-00"


def pc_from_metadata(metadata: dict[str, str], path: Path) -> str:
    return canonical_pc(metadata.get("N° Resolucion") or path.stem)


def cod_arbol(pc: str, codigo: str) -> str:
    return f"{pc}-{codigo}"


def make_troza_id(pc: str, codigo_troza: str) -> str:
    return f"{pc}-{codigo_troza.replace('/', '-')}"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value).strip()


def to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


def distance(a: float, b: float, c: float, d: float) -> float:
    if not all([a, b, c, d]):
        return 0.0
    return round(((a - c) ** 2 + (b - d) ** 2) ** 0.5, 2)


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    if not text:
        return "2025-07-22"
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10]


def parse_slash_date(value: str) -> str:
    return datetime.strptime(value, "%d/%m/%Y").date().isoformat()


def sha256_json(payload: Any) -> str:
    return hashlib.sha256(json.dumps(payload, sort_keys=True, default=str, ensure_ascii=False).encode("utf-8")).hexdigest()


if __name__ == "__main__":
    main()
